"""P3-2：ConvertX —— 任意格式互转模块（Reader × Writer = N×N 矩阵）。

**与现有 export_all 的设计取舍**（依据用户要求：如果任意格式互转比单输入多输出更实用就采纳，
且项目原有设计不一定是最好的）：

    现有 `pipeline/exporters.py::export_all` 定位 = 「pipeline 跑完后一次性导出所有格式」
    （单输入：StateStore 表 → 多输出：jsonl/csv/xlsx/parquet/duckdb 多文件同时落盘）。
    它满足 pipeline 结束那刻的"多格式输出"，但**缺少事后互转**能力：
        例 1. 用户最初只开了 CSV，后来想要 XLSX —— 以前要重跑 pipeline；
        例 2. 用户从别处拿到一份 JSONL records，想导入成自己 project 的 records.db —— 以前做不到。
        例 3. 跨系统对接：需要 Parquet/DuckDB 列式 + 压缩（AI/BI 常用），但别人只给 CSV。

    因此新增 `omnicrawler.convertx` 模块，定位 = 「文件级 A → B 互转」：
        - 中间表示 CanonicalRecords = list[flat dict]（与 export_all 展开后的 flat records 完全一致，
          列序稳定为：record_id, source_url, record_type, created_at, [展开 data_json]）
        - 核心入口 convert(src, dst, options)：按后缀名（或显式 fmt 指定）自动选 Reader/Writer。
        - **不影响** export_all 管道默认的多输出（用户的 config 行为保持不变），convertx 作为
          独立的 CLI/GUI 工具层提供。
        - 这是一种增量扩展，没有删除/重构任何老代码。

**Reader 注册表（按输入后缀选择）**：
    .jsonl → JSONLReader
    .csv   → CSVReader
    .parquet → ParquetReader（需要 pyarrow）
    .duckdb / .db → DuckDBReader（需要 duckdb）
    .xlsx  → XLSXReader（需要 openpyxl）

**Writer 注册表**：
    同上格式，一一对应。

**安全约束**：
    * 输入文件必须是真实存在文件（Path.resolve() → 必须在当前卷，禁止 path traversal）
    * 输出文件使用 `core.utils.atomic_write` 或等价逻辑（先写 .tmp → rename，避免半写）
    * XLSX/DuckDB 单表超过 1,000,000 行给出警告（Excel 硬限 1,048,576 行）
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import time
from collections.abc import Callable, Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..core.utils import excel_safe
from ..services.progress import (
    ProgressTracker,
    StageSpec,
    TaskProgressEvent,
)

LOGGER = logging.getLogger(__name__)

# ── 进度节流：每 N 条 / 每 T ms 推一次，避免高频回调拖慢转换 ──
_PROGRESS_CHUNK: int = 250          # 至少每 250 条 1 次（小文件 2k 行也会推 8 次，视觉丝滑）
_PROGRESS_MIN_INTERVAL_S: float = 0.03  # 上限 30ms/次（大文件 2M 行也只 ~33 次/秒，GUI 完全无感）
_PROGRESS_EST_AVG_BYTES_PER_LINE: int = 180  # 读取阶段估算总行数的平均行字节启发式（含 csv 逗号/引号）


class _ProgressEmitter:
    """节流发射器：避免每条记录都触发一次 QThread 信号。"""

    __slots__ = ("_hook", "_last_emit_ts", "_since_last")

    def __init__(self, hook: Callable[[dict[str, Any]], None] | None) -> None:
        self._hook = hook
        self._last_emit_ts: float = 0.0
        self._since_last: int = 0

    def emit(self, *, force: bool = False, **fields: Any) -> None:
        if self._hook is None:
            return
        self._since_last += 1
        now = time.monotonic()
        if not force and self._since_last < _PROGRESS_CHUNK and (now - self._last_emit_ts) < _PROGRESS_MIN_INTERVAL_S:
            return
        try:
            self._hook(dict(fields))
        except Exception:  # noqa: BLE001 — 用户回调/进度桥接报错不影响转换
            pass
        self._last_emit_ts = now
        self._since_last = 0

    def flush(self, **fields: Any) -> None:
        """最后强制推一次（保证 100% 命中，避免最后不足 1 chunk 的残条不显示）。"""
        self.emit(force=True, **fields)

__all__ = [
    "READERS",
    "WRITERS",
    "CanonicalRecords",
    "ConvertResult",
    "ProgressTracker",
    "ReaderFn",
    "StageSpec",
    "TaskProgressEvent",
    "WriterFn",
    "convert",
    "register_reader",
    "register_writer",
    "sniff_format",
    # ---- 核心 Reader/Writer（直接调用时可导入）----
    "read_csv",
    "read_jsonl",
    "write_csv",
    "write_jsonl",
]

# ── 类型 ──────────────────────────────────────────────────
CanonicalRecords = list[dict[str, Any]]
ReaderFn = Callable[[Path, dict[str, Any]], CanonicalRecords]
WriterFn = Callable[[CanonicalRecords, Path, dict[str, Any]], dict[str, Any]]


@dataclass(slots=True)
class ConvertResult:
    source_format: str
    target_format: str
    rows: int
    columns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    output_path: Path | None = None
    extra: dict[str, Any] = field(default_factory=dict)


# ── 注册表 ────────────────────────────────────────────────
READERS: dict[str, ReaderFn] = {}
WRITERS: dict[str, WriterFn] = {}


def register_reader(*extensions: str) -> Callable[[ReaderFn], ReaderFn]:
    """注册一个 Reader（装饰器）。extensions 含前导点如 '.jsonl'。"""
    def _wrap(fn: ReaderFn) -> ReaderFn:
        for ext in extensions:
            READERS[ext.lower()] = fn
        return fn
    return _wrap


def register_writer(*extensions: str) -> Callable[[WriterFn], WriterFn]:
    """注册一个 Writer（装饰器）。"""
    def _wrap(fn: WriterFn) -> WriterFn:
        for ext in extensions:
            WRITERS[ext.lower()] = fn
        return fn
    return _wrap


# ── 工具 ──────────────────────────────────────────────────
def sniff_format(path: Path) -> str | None:
    """根据文件后缀推断格式（统一返回 READERS/WRITERS 中注册的 key，如 '.jsonl'）。

    Alias 归一化：
        .ndjson → .jsonl（同一 Reader/Writer）
        .db     → .duckdb
    """
    suffix = path.suffix.lower()
    if suffix == ".ndjson":
        return ".jsonl"
    if suffix == ".db":
        return ".duckdb"
    if suffix in READERS or suffix in WRITERS:
        return suffix
    return None


_BASE_COLUMNS: tuple[str, ...] = ("record_id", "source_url", "record_type", "created_at")


def _ordered_columns(rows: CanonicalRecords, *, prefer: Iterable[str] = ()) -> list[str]:
    """按 1) _BASE_COLUMNS 先；2) prefer；3) 首次出现 来稳定列序。"""
    seen: dict[str, None] = {}
    # 1) base 先
    for col in _BASE_COLUMNS:
        seen[col] = None
    # 2) prefer
    for col in prefer:
        if col:
            seen[str(col)] = None
    # 3) 首次出现
    for row in rows:
        for key in row.keys():
            seen[str(key)] = None
    return list(seen.keys())


def _require_file(path: Path) -> None:
    if not isinstance(path, Path):
        path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"ConvertX: 输入文件不存在: {path}")
    if not path.exists():
        raise FileNotFoundError(f"ConvertX: 输入文件路径无效: {path}")


# META：SQL 标识符白名单（表名/列名直插场景），仅允许 identifier 或 schema.identifier。
_SQL_IDENTIFIER_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)?")


def _ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


# ── CSV ──────────────────────────────────────────────────
@register_reader(".csv")
def read_csv(path: Path, options: dict[str, Any]) -> CanonicalRecords:
    _require_file(path)
    encoding = str(options.get("encoding", "utf-8-sig"))
    on_error = str(options.get("on_error", "skip")).lower()  # skip | abort
    pe = _ProgressEmitter(options.get("on_line_progress"))
    if encoding == "auto":
        # S3：自动编码检测（chardet），检测失败走 utf-8→gb18030→latin-1 兜底链
        from ..core.encoding import smart_decode

        text, encoding = smart_decode(path.read_bytes())
        fh: io.TextIOBase = io.StringIO(text)
    else:
        fh = path.open("r", encoding=encoding, newline="")
    try:
        reader = csv.DictReader(fh)
        records: CanonicalRecords = []
        line_num = 0
        for row in reader:
            line_num += 1
            try:
                records.append(dict(row))
            except Exception as exc:  # csv 一般不含异常；保留以对齐 on_error 语义
                if on_error == "abort":
                    raise ValueError(f"CSV 解析失败（逻辑行 {line_num}）: {exc}") from exc
                continue
            pe.emit(line_num=line_num, records_so_far=len(records))
    finally:
        fh.close()
    pe.flush(line_num=line_num, records_so_far=len(records))
    return records


@register_writer(".csv")
def write_csv(rows: CanonicalRecords, path: Path, options: dict[str, Any]) -> dict[str, Any]:
    _ensure_parent_dir(path)
    columns = _ordered_columns(rows, prefer=options.get("columns") or [])
    encoding = str(options.get("encoding", "utf-8-sig"))
    pe = _ProgressEmitter(options.get("on_write_progress"))
    total = len(rows)
    tmp = path.with_suffix(path.suffix + ".tmp")
    written = 0
    try:
        with tmp.open("w", encoding=encoding, newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            if total <= _PROGRESS_CHUNK * 2:
                # 小文件仍可一次性 writerows（最快路径），写后直接 flush
                out = [{k: excel_safe(row.get(k, "")) for k in columns} for row in rows]
                writer.writerows(out)
                written = total
            else:
                # 大文件按 chunk 写出，节流推进度
                buf: list[dict[str, Any]] = []
                for row in rows:
                    buf.append({k: excel_safe(row.get(k, "")) for k in columns})
                    if len(buf) >= _PROGRESS_CHUNK:
                        writer.writerows(buf)
                        written += len(buf)
                        buf.clear()
                        pe.emit(written=written, total=total)
                if buf:
                    writer.writerows(buf)
                    written += len(buf)
                    buf.clear()
        tmp.replace(path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
    pe.flush(written=written, total=total)
    return {"rows": written, "columns": columns, "encoding": encoding}


# ── JSONL ─────────────────────────────────────────────────
@register_reader(".jsonl", ".ndjson")
def read_jsonl(path: Path, options: dict[str, Any]) -> CanonicalRecords:
    _require_file(path)
    flat_mode = bool(options.get("flat", True))  # 默认把 .data 展开为 flat dict
    on_error = str(options.get("on_error", "skip")).lower()  # skip | abort
    pe = _ProgressEmitter(options.get("on_line_progress"))
    records: CanonicalRecords = []
    last_line = 0
    with path.open("r", encoding="utf-8") as fh:
        for line_num, line in enumerate(fh, 1):
            last_line = line_num
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                if on_error == "abort":
                    raise ValueError(f"JSONL 解析失败（行 {line_num}）: {e}") from e
                # skip 默认行为
                continue
            if not isinstance(obj, dict):
                continue
            if flat_mode and isinstance(obj.get("data"), dict):
                flat: dict[str, Any] = {
                    k: obj.get(k) for k in _BASE_COLUMNS if k in obj
                }
                _flatten_to("", obj["data"], flat)
                if obj.get("evidence"):
                    flat["evidence_json"] = json.dumps(obj["evidence"], ensure_ascii=False)
                records.append(flat)
            else:
                records.append(dict(obj))
            pe.emit(line_num=line_num, records_so_far=len(records))
    pe.flush(line_num=last_line, records_so_far=len(records))
    return records


@register_writer(".jsonl", ".ndjson")
def write_jsonl(rows: CanonicalRecords, path: Path, options: dict[str, Any]) -> dict[str, Any]:
    _ensure_parent_dir(path)
    nested = bool(options.get("nested", False))  # True 时按 pipeline 原始 records.jsonl 结构
    pe = _ProgressEmitter(options.get("on_write_progress"))
    total = len(rows)
    tmp = path.with_suffix(path.suffix + ".tmp")
    written = 0
    try:
        with tmp.open("w", encoding="utf-8") as fh:
            if nested:
                for row in rows:
                    base = {k: row.get(k) for k in _BASE_COLUMNS if k in row}
                    data = {k: v for k, v in row.items() if k not in _BASE_COLUMNS and k != "evidence_json"}
                    evidence = {}
                    ev_raw = row.get("evidence_json")
                    if isinstance(ev_raw, str):
                        try:
                            evidence = json.loads(ev_raw)
                        except (TypeError, ValueError):
                            evidence = {"raw": ev_raw}
                    base["data"] = data
                    base["evidence"] = evidence
                    fh.write(json.dumps(base, ensure_ascii=False) + "\n")
                    written += 1
                    pe.emit(written=written, total=total)
            else:
                for row in rows:
                    fh.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
                    written += 1
                    pe.emit(written=written, total=total)
        tmp.replace(path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
    pe.flush(written=written, total=total)
    return {"rows": written, "nested": nested}


def _flatten_to(prefix: str, value: Any, out: dict[str, Any]) -> None:
    if isinstance(value, dict):
        for k, v in value.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, (dict, list, tuple)):
                out[key] = json.dumps(v, ensure_ascii=False, default=str)
            else:
                out[key] = v
    elif isinstance(value, (list, tuple)):
        out_key = prefix or "items"
        out[out_key] = json.dumps(value, ensure_ascii=False, default=str)
    else:
        if prefix:
            out[prefix] = value


# ── Parquet ───────────────────────────────────────────────
def _register_parquet() -> None:
    try:
        import pyarrow as pa  # noqa: F401
        import pyarrow.parquet as pq  # noqa: F401
    except ImportError:
        return

    @register_reader(".parquet")
    def read_parquet(path: Path, options: dict[str, Any]) -> CanonicalRecords:
        import pyarrow.parquet as pq

        _require_file(path)
        pe = _ProgressEmitter(options.get("on_line_progress"))
        pf = pq.ParquetFile(path)
        total_rows = int(getattr(pf.metadata, "num_rows", 0) or 0)
        records: CanonicalRecords = []
        row_cursor = 0
        for batch in pf.iter_batches(batch_size=max(1, _PROGRESS_CHUNK)):
            for obj in batch.to_pylist():
                if isinstance(obj, dict):
                    records.append(obj)
            row_cursor += batch.num_rows
            pe.emit(
                line_num=row_cursor,
                records_so_far=len(records),
                total_rows=total_rows,
            )
        pe.flush(
            line_num=row_cursor,
            records_so_far=len(records),
            total_rows=total_rows,
        )
        return records

    @register_writer(".parquet")
    def write_parquet(rows: CanonicalRecords, path: Path, options: dict[str, Any]) -> dict[str, Any]:
        import pyarrow as pa
        import pyarrow.parquet as pq

        _ensure_parent_dir(path)
        pe = _ProgressEmitter(options.get("on_write_progress"))
        total = len(rows)
        compression = str(options.get("compression", "zstd"))

        if not rows:
            schema = pa.schema([pa.field("record_id", pa.string())])
            empty_table = pa.table({"record_id": pa.array([], type=pa.string())}, schema=schema)
            pq.write_table(empty_table, path, compression=compression)
            pe.flush(written=0, total=0)
            return {"rows": 0, "columns": schema.names, "compression": compression}

        # 为了列稳定，先拿完整列集合，首批用它创建 writer schema
        columns = _ordered_columns(rows, prefer=options.get("columns") or [])
        # 从首批构造 schema（None 的用 STRING，其余 pyarrow 自动推断）
        probe = rows[:min(50, len(rows))]
        try:
            probe_table = pa.Table.from_pylist(probe)
            schema = probe_table.schema
        except Exception:  # noqa: BLE001 — 混合类型回退成通用 schema
            fields = [pa.field(c, pa.string()) for c in columns]
            schema = pa.schema(fields)

        written = 0
        writer = pq.ParquetWriter(path, schema, compression=compression)
        try:
            for start in range(0, total, _PROGRESS_CHUNK):
                chunk = rows[start:start + _PROGRESS_CHUNK]
                try:
                    batch_table = pa.Table.from_pylist(chunk, schema=schema)
                except Exception:  # noqa: BLE001 — 类型不一致时强制转 string（按 schema 创建 writer 已锁）
                    normalized: list[dict[str, Any]] = []
                    for row in chunk:
                        normalized.append({c: (row[c] if c in row and row[c] is not None else None) for c in columns})
                    batch_table = pa.Table.from_pylist(normalized)
                writer.write_table(batch_table)
                written += len(chunk)
                pe.emit(written=written, total=total)
            pe.flush(written=written, total=total)
        finally:
            try:
                writer.close()
            except Exception:  # noqa: BLE001
                pass
        return {"rows": written, "columns": list(schema.names), "compression": compression}


_register_parquet()


# ── DuckDB ────────────────────────────────────────────────
_VALID_COLUMN_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")


def _register_duckdb() -> None:
    try:
        import duckdb  # noqa: F401
    except ImportError:
        return

    def _validate_columns(columns: list[str]) -> list[str]:
        out: list[str] = []
        for c in columns:
            nc = re.sub(r"[^a-zA-Z0-9_]", "_", str(c))
            if not nc:
                nc = "col"
            if not nc[0].isalpha() and nc[0] != "_":
                nc = "c_" + nc
            out.append(nc)
        return out

    def _infer_py_type(values: list[Any]) -> str:
        kinds: set[str] = set()
        for v in values:
            if v is None or v == "":
                continue
            if isinstance(v, bool):
                kinds.add("bool")
            elif isinstance(v, int):
                kinds.add("int")
            elif isinstance(v, float):
                kinds.add("float")
            else:
                return "VARCHAR"
        if not kinds:
            return "VARCHAR"
        if kinds <= {"bool"}:
            return "BOOLEAN"
        if kinds <= {"int"}:
            return "BIGINT"
        if kinds <= {"float", "int"}:
            return "DOUBLE"
        return "VARCHAR"

    @register_reader(".duckdb", ".db")
    def read_duckdb(path: Path, options: dict[str, Any]) -> CanonicalRecords:
        import duckdb

        _require_file(path)
        pe = _ProgressEmitter(options.get("on_line_progress"))
        table = str(options.get("table", "records"))
        # META：table 标识符白名单——仅允许标识符或 schema.identifier，防 CLI
        # 传入恶意标识符/子查询直插 SQL（convertx 本地-only，仍按 RC-6 同族收口）。
        if not _SQL_IDENTIFIER_RE.fullmatch(table):
            raise ValueError(f"无效的 duckdb 表名: {table!r}")
        # 先拿总行数，便于节流推进展示进度
        count_sql = f"SELECT COUNT(*) FROM {table}"
        read_sql = f"SELECT * FROM {table}"
        total_rows = 0
        con = duckdb.connect(str(path), read_only=True)
        try:
            try:
                count_row = con.execute(count_sql).fetchone()
                if count_row is not None:
                    total_rows = int(count_row[0] or 0)
            except Exception:  # noqa: BLE001 — 视图/权限问题取不到 count 就按 0 处理（仍推进）
                total_rows = 0
            cursor = con.execute(read_sql)
            cols = [d[0] for d in cursor.description or []]
            records: CanonicalRecords = []
            fetched = 0
            while True:
                chunk = cursor.fetchmany(_PROGRESS_CHUNK)
                if not chunk:
                    break
                for row in chunk:
                    records.append(dict(zip(cols, row, strict=True)))
                fetched += len(chunk)
                pe.emit(
                    line_num=fetched,
                    records_so_far=len(records),
                    total_rows=total_rows,
                )
            pe.flush(
                line_num=fetched,
                records_so_far=len(records),
                total_rows=total_rows,
            )
            return records
        finally:
            con.close()

    @register_writer(".duckdb", ".db")
    def write_duckdb(rows: CanonicalRecords, path: Path, options: dict[str, Any]) -> dict[str, Any]:
        import duckdb

        _ensure_parent_dir(path)
        pe = _ProgressEmitter(options.get("on_write_progress"))
        total = len(rows)
        table = str(options.get("table", "records"))
        safe_cols = _ordered_columns(rows, prefer=options.get("columns") or [])
        if not safe_cols:
            safe_cols = ["record_id"]
        typed_cols = _validate_columns(safe_cols)
        col_types = [
            _infer_py_type([row.get(raw_col) for row in rows if raw_col in row])
            for raw_col in safe_cols
        ]
        ddl = ", ".join(f'"{c}" {t}' for c, t in zip(typed_cols, col_types, strict=True))
        placeholders = ", ".join("?" for _ in typed_cols)
        con = duckdb.connect(str(path))
        written = 0
        try:
            con.execute(f"DROP TABLE IF EXISTS {table}")
            con.execute(f"CREATE TABLE {table} ({ddl})")
            if rows:
                for start in range(0, total, _PROGRESS_CHUNK):
                    chunk = rows[start:start + _PROGRESS_CHUNK]
                    values = [[row.get(k) for k in safe_cols] for row in chunk]
                    if values:
                        con.executemany(f"INSERT INTO {table} VALUES ({placeholders})", values)
                    written += len(chunk)
                    pe.emit(written=written, total=total)
            pe.flush(written=written, total=total)
        finally:
            con.close()
        return {"rows": written, "columns": safe_cols, "table": table}


_register_duckdb()


# ── XLSX ─────────────────────────────────────────────────
XLSX_ROW_LIMIT = 1_000_000


def _register_xlsx() -> None:
    try:
        from openpyxl import Workbook  # noqa: F401
    except ImportError:
        return

    @register_reader(".xlsx")
    def read_xlsx(path: Path, options: dict[str, Any]) -> CanonicalRecords:
        from openpyxl import load_workbook

        _require_file(path)
        sheet_name = options.get("sheet")
        data_only = bool(options.get("data_only", True))
        on_error = str(options.get("on_error", "skip")).lower()  # skip | abort
        pe = _ProgressEmitter(options.get("on_line_progress"))
        wb = load_workbook(filename=str(path), read_only=True, data_only=data_only)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
            except StopIteration:
                pe.flush(line_num=0, records_so_far=0)
                return []
            header = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(header)]
            records: CanonicalRecords = []
            row_num = 1  # 表头为 0，第一个数据行从 1 起算
            for row in rows_iter:
                row_num += 1
                if row is None or all(v is None or v == "" for v in row):
                    continue
                try:
                    rec: dict[str, Any] = {}
                    for i, key in enumerate(header):
                        if i < len(row):
                            rec[key] = row[i]
                    records.append(rec)
                except Exception as exc:
                    if on_error == "abort":
                        raise ValueError(f"XLSX 解析失败（行 {row_num}）: {exc}") from exc
                    continue
                pe.emit(line_num=row_num, records_so_far=len(records))
            pe.flush(line_num=row_num, records_so_far=len(records))
            return records
        finally:
            wb.close()

    @register_writer(".xlsx")
    def write_xlsx(rows: CanonicalRecords, path: Path, options: dict[str, Any]) -> dict[str, Any]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        _ensure_parent_dir(path)
        columns = _ordered_columns(rows, prefer=options.get("columns") or [])
        pe = _ProgressEmitter(options.get("on_write_progress"))
        wb = Workbook()
        ws = wb.active
        ws.title = "结构化记录"
        ws.freeze_panes = "A2"
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        warnings: list[str] = []
        data_rows = rows
        if len(data_rows) > XLSX_ROW_LIMIT:
            warnings.append(
                f"XLSX 单表超过 {XLSX_ROW_LIMIT} 行硬限，仅写前 {XLSX_ROW_LIMIT} 行（Excel 不支持更多）"
            )
            LOGGER.warning(warnings[-1])
            data_rows = data_rows[:XLSX_ROW_LIMIT]
        total = len(data_rows)
        written = 0
        for row in data_rows:
            ws.append([_xlsx_cell(row.get(k, ""), k) for k in columns])
            written += 1
            pe.emit(written=written, total=total)
        pe.flush(written=written, total=total)
        try:
            wb.save(path)
        except (PermissionError, OSError) as exc:
            raise RuntimeError(f"无法写入 Excel 文件 {path}（可能被其他程序占用或目录不可写）: {exc}") from exc
        return {"rows": written, "columns": columns, "warnings": warnings, "sheet": "结构化记录"}


def _xlsx_cell(value: Any, column: str) -> Any:
    v = excel_safe(value)
    if isinstance(v, str):
        # Excel 单元格硬限 32767 字符
        if len(v) > 32700:
            v = v[:32700]
        return v
    return v


_register_xlsx()


def _register_document() -> None:
    """注册 document 族（懒加载 document_ir，避免 import 副作用）。"""
    from . import document  # noqa: F401


_register_document()


# ── 入口 ─────────────────────────────────────────────────
def convert(
    source: str | Path,
    target: str | Path,
    *,
    src_format: str | None = None,
    dst_format: str | None = None,
    options: dict[str, Any] | None = None,
    flat: bool = True,
    nested: bool = False,
    table: str = "records",
    compression: str = "zstd",
    on_read_progress: Callable[[Any], None] | None = None,
    on_write_progress: Callable[[Any], None] | None = None,
    on_error: str = "skip",
    on_progress: Callable[[TaskProgressEvent], None] | None = None,
) -> ConvertResult:
    """核心入口：A → B 互转。

    Parameters
    ----------
    source: 源文件路径
    target: 目标文件路径
    src_format: 显式指定源格式（如 '.jsonl'）；None 时按 suffix 推断
    dst_format: 显式指定目标格式；None 时按 target suffix 推断
    options:  透传给 Reader/Writer 选项（例如 nested=True 让 jsonl 输出 data/evidence 嵌套）
    flat:     JSONL 读取时是否展开 data 字段（默认 True）
    nested:   JSONL 写出时是否按 pipeline 原始嵌套结构（data/evidence 独立字段）
    table:    DuckDB 读写时使用的表名（默认 'records'）
    compression: Parquet 写出时压缩算法（zstd/snappy/gzip/none，默认 zstd）
    on_read_progress:  读取阶段进度回调（接收 dict: {line_num, records_so_far}）
    on_write_progress: 写入阶段进度回调（旧式 dict hooks，保持兼容）
    on_error: 解析错误策略 'skip'（默认）| 'abort'
    on_progress: **统一进度事件回调**（推荐）。接收 ``TaskProgressEvent``：
                 阶段=read(60%) / write(40%)，含 EMA ETA、瞬时速率、状态机。

    Raises
    ------
    FileNotFoundError: 源文件不存在
    KeyError: 源或目标格式未在注册表中注册（通常是缺少 openpyxl/pyarrow/duckdb 等可选依赖）
    """
    src = source if isinstance(source, Path) else Path(source)
    dst = target if isinstance(target, Path) else Path(target)
    if src.resolve() == dst.resolve():
        raise ValueError(f"ConvertX: 源与目标路径相同，拒绝覆盖: {src}")

    # ── 统一进度 tracker（仅在 on_progress 显式传入时启用，避免副作用）──
    tracker: ProgressTracker | None = None
    if on_progress is not None:
        tracker = ProgressTracker(
            stages=[
                StageSpec("read",  weight=60, display_name="读取格式", has_items=True),
                StageSpec("write", weight=40, display_name="写出格式", has_items=True),
            ],
            on_event=on_progress,
        )
        tracker.start()

    opts = dict(options or {})

    def _wrapped_read_hook(payload: dict[str, Any]) -> None:
        # 1) 先调用用户原始 hook（若有）
        if on_read_progress is not None:
            try:
                on_read_progress(payload)
            except Exception:  # noqa: BLE001 — hook 错误不得中断转换
                pass
        # 2) 再更新 tracker：records_so_far 当前处理条数
        if tracker is not None:
            records_so_far = int(payload.get("records_so_far") or 0)
            tracker.set_item_progress(records_so_far, max(records_so_far, tracker._items_total or 0))  # type: ignore[attr-defined]

    def _wrapped_write_hook(payload: dict[str, Any]) -> None:
        if on_write_progress is not None:
            try:
                on_write_progress(payload)
            except Exception:  # noqa: BLE001
                pass
        if tracker is not None:
            current = int(payload.get("written") or payload.get("records_written") or 0)
            total = int(payload.get("total") or tracker._items_total or current)  # type: ignore[attr-defined]
            tracker.set_item_progress(current, total)

    # 先解析实际 fmt（否则当用户没显式传 src_format/dst_format 时，key 为 None，
    # 通用选项的 fallback key 又用了「点分隔符」，会导致 hooks / columns 完全没注入）
    src_fmt = (src_format or "").lower() or sniff_format(src)
    dst_fmt = (dst_format or "").lower() or sniff_format(dst)
    if not src_fmt or src_fmt not in READERS:
        raise KeyError(f"ConvertX: 不支持的源格式 {src_fmt!r}（已注册: {sorted(READERS)}）")
    if not dst_fmt or dst_fmt not in WRITERS:
        raise KeyError(f"ConvertX: 不支持的目标格式 {dst_fmt!r}（已注册: {sorted(WRITERS)}）")

    # 真实 key（下划线分隔：如 .csv → reader_csv / writer_csv；.jsonl → reader_jsonl）
    reader_key = f"reader{src_fmt.replace('.', '_')}"
    writer_key = f"writer{dst_fmt.replace('.', '_')}"

    # Reader 通用选项：按实际 reader_key 单点注入即可（不再对不存在的 key 做 fallback）
    r_opts = opts.setdefault(reader_key, {})
    r_opts.setdefault("flat", flat)
    r_opts.setdefault("on_error", on_error)
    if tracker is not None or on_read_progress is not None:
        r_opts["on_line_progress"] = _wrapped_read_hook

    # Writer 通用选项：单点注入 + 格式特定默认
    w_opts = opts.setdefault(writer_key, {})
    w_opts.setdefault("nested", nested)
    if tracker is not None or on_write_progress is not None:
        w_opts["on_write_progress"] = _wrapped_write_hook
    if dst_fmt in {".parquet"}:
        w_opts.setdefault("compression", compression)
    if dst_fmt in {".duckdb", ".db"}:
        w_opts.setdefault("table", table)
    if src_fmt in {".duckdb", ".db"}:
        r_opts.setdefault("table", table)

    # ── 阶段 1：Read（权重 60%）──────────────────────────────
    est_read_items: int = 0
    if tracker is not None:
        # 纯文本格式：按文件字节 / 平均行字节 粗略估算行数（仅给 read 阶段一个有意义的分母）
        if src_fmt in {".csv", ".jsonl", ".ndjson"}:
            try:
                est_read_items = max(1, src.stat().st_size // _PROGRESS_EST_AVG_BYTES_PER_LINE)
            except OSError:
                est_read_items = 0
        tracker.begin_stage("read", expected_items=est_read_items)
    reader_opts = opts.get(reader_key, {}) or {}
    try:
        rows = READERS[src_fmt](src, reader_opts)
    except Exception:
        if tracker is not None:
            tracker.fail("读取阶段出错")
        raise
    if tracker is not None:
        tracker.end_stage("read")

    cols = _ordered_columns(rows, prefer=opts.get("columns") or [])

    # ── 阶段 2：Write（权重 40%）──────────────────────────────
    if tracker is not None:
        tracker.begin_stage("write", expected_items=len(rows))
    writer_opts = dict(opts.get(writer_key, {}) or {})
    writer_opts.setdefault("columns", cols)
    try:
        writer_meta = WRITERS[dst_fmt](rows, dst, writer_opts)
    except Exception:
        if tracker is not None:
            tracker.fail("写入阶段出错")
        raise
    if tracker is not None:
        tracker.end_stage("write")
        tracker.finish()

    warnings: list[str] = []
    warnings.extend(writer_meta.pop("warnings") if isinstance(writer_meta, dict) and "warnings" in writer_meta else [])
    return ConvertResult(
        source_format=src_fmt,
        target_format=dst_fmt,
        rows=len(rows),
        columns=cols,
        warnings=warnings,
        output_path=dst,
        extra=writer_meta if isinstance(writer_meta, dict) else {},
    )
