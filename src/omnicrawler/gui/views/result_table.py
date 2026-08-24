"""结果表格视图。

QTableView + 流式 CSV Model，支持分页、排序、Excel 导出。
CSV 索引（行计数）与 JSONL 证据查找均通过后台线程执行，避免大文件阻塞 UI。
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

from PySide6.QtCore import (
    QAbstractTableModel,
    QEasingCurve,
    QModelIndex,
    QPropertyAnimation,
    QRegularExpression,
    QSortFilterProxyModel,
    Qt,
    QThread,
    Signal,
)
from PySide6.QtWidgets import (
    QFileDialog,
    QGraphicsOpacityEffect,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QProgressDialog,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from ...core.utils import excel_safe
from ..async_workers import CsvIndexWorker, JsonlSearchWorker
from ..i18n import _

ROWS_PER_PAGE = 1000

_logger = logging.getLogger(__name__)


class CsvStreamModel(QAbstractTableModel):
    """流式 CSV 数据模型。

    按需加载 CSV 行，避免启动时扫描全文。
    表头与行计数通过 CsvIndexWorker 异步获取，避免大文件阻塞 UI。
    """

    indexing_started = Signal()
    indexing_finished = Signal(int)  # total_rows
    indexing_failed = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._filepath: Path | None = None
        self._headers: list[str] = []
        self._rows: list[list[str]] = []
        self._total_rows: int = 0
        self._file_size: int = 0
        self._avg_line_size: int = 200  # 估计平均行长
        self._line_offsets: list[int] = []
        self._index_worker: CsvIndexWorker | None = None
        self.last_error: str | None = None  # A17：最近一次翻页/加载错误（视图层展示）

    def load_file_async(self, filepath: Path) -> bool:
        """异步加载 CSV 文件。

        立即返回 True 表示已派发索引任务。
        表头和行计数通过 indexing_finished 信号通知；失败通过 indexing_failed。
        """
        # 取消尚未完成的旧索引任务
        if self._index_worker is not None and self._index_worker.isRunning():
            self._index_worker.requestInterruption()
            self._index_worker.quit()
            self._index_worker.wait(2000)

        self.beginResetModel()
        self._filepath = filepath
        try:
            self._file_size = filepath.stat().st_size
        except OSError:
            self._file_size = 0
        self._rows = []
        self._line_offsets = []
        self._headers = []
        self._total_rows = 0
        self.endResetModel()

        self.indexing_started.emit()

        self._index_worker = CsvIndexWorker(filepath, parent=self)
        self._index_worker.finished_indexing.connect(self._on_indexed)
        self._index_worker.failed.connect(self._on_index_failed)
        self._index_worker.finished.connect(self._on_index_worker_finished)
        self._index_worker.start()
        return True

    def _on_indexed(self, headers: list, total_rows: int, file_size: float) -> None:
        """索引完成回调（主线程）。"""
        self.beginResetModel()
        self._headers = headers
        self._total_rows = total_rows
        self._file_size = int(file_size)
        self._line_offsets = list(range(0, total_rows, ROWS_PER_PAGE)) or [0]
        self._load_page(0)
        self.endResetModel()
        self.indexing_finished.emit(total_rows)

    def _on_index_failed(self, message: str) -> None:
        """索引失败回调（主线程）。"""
        self.indexing_failed.emit(message)

    def _on_index_worker_finished(self) -> None:
        """线程结束后清理引用。"""
        sender = self.sender()
        if sender is self._index_worker:
            self._index_worker = None

    def load_file(self, filepath: Path) -> bool:
        """同步加载 CSV 文件（仅读表头和建立偏移索引）。

        保留同步入口以兼容已有调用；新代码应优先使用 load_file_async。
        """
        self.beginResetModel()
        try:
            self._filepath = filepath
            self._file_size = filepath.stat().st_size
            self._rows = []
            self._line_offsets = []

            with open(filepath, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                first_row = next(reader, None)
                if first_row is None:
                    self.endResetModel()
                    return False
                self._headers = [str(h).strip() for h in first_row]

                # csv.reader 使用文本缓冲后不允许可靠 tell()。只记录每页起始行号
                # （B9：不再截断——_total_rows 完整计数，_line_offsets 内存 O(页数)）。
                row_count = 0
                for _ in reader:
                    if row_count % ROWS_PER_PAGE == 0:
                        self._line_offsets.append(row_count)
                    row_count += 1
                self._total_rows = row_count

            if self._total_rows == 0:
                self._total_rows = 0
            self._line_offsets = self._line_offsets or [0]

            self._load_page(0)
            self.endResetModel()
            return True
        except Exception:
            self.endResetModel()
            return False

    def _load_page(self, page: int) -> None:
        """加载指定页的数据。"""
        if self._filepath is None:
            return

        start = page * ROWS_PER_PAGE
        if self._total_rows > 0 and start >= self._total_rows:
            self._rows = []
            return

        self._rows = []
        try:
            with open(self._filepath, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                next(reader, None)  # skip header
                for _ in range(start):
                    if next(reader, None) is None:
                        break

                count = 0
                for row in reader:
                    self._rows.append([str(c).strip() for c in row])
                    count += 1
                    if count >= ROWS_PER_PAGE:
                        break
        except Exception as exc:
            _logger.warning("Failed to load page %d: %s", page, exc)
            self.last_error = str(exc)

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]  # noqa: B008
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]  # noqa: B008
        return len(self._headers)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            row = index.row()
            col = index.column()
            try:
                return self._rows[row][col]
            except (IndexError, TypeError):
                return ""
        return None

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            if section < len(self._headers):
                return self._headers[section]
        return super().headerData(section, orientation, role)

    def canFetchMore(self, index: QModelIndex) -> bool:  # type: ignore[override]
        return False

    @property
    def total_rows(self) -> int:
        return self._total_rows

    @property
    def total_pages(self) -> int:
        if self._total_rows == 0:
            return 0
        return max(1, (self._total_rows + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)

    @property
    def headers(self) -> list[str]:
        return list(self._headers)

    def go_to_page(self, page: int) -> None:
        """跳转到指定页码。"""
        self.last_error = None
        total = self.total_pages
        if total == 0:
            return
        page = max(0, min(page, total - 1))
        self.beginResetModel()
        self._load_page(page)
        self.endResetModel()


class ExportThread(QThread):
    """Excel 导出后台线程。

    P2-4：通过 ProgressTracker 阶段权重 + 子项计数驱动进度，
    保持旧式 progress(int) 信号发出（旧消费者不受影响）。
    """

    progress = Signal(int)
    finished_signal = Signal(bool, str)
    unified_progress = Signal(object)  # P2-4：TaskProgressEvent

    def __init__(self, filepath: Path, output_path: Path) -> None:
        super().__init__()
        self._filepath = filepath
        self._output_path = output_path

    def run(self) -> None:
        try:
            from omnicrawler.services.progress import ProgressTracker, StageSpec

            if self.isInterruptionRequested():
                return
            import openpyxl

            # P2-4：只做"估算总行数→三阶段进度"接入，保持流式读写，不将 CSV 全量加载进内存
            total_rows = 0
            try:
                with open(self._filepath, encoding="utf-8-sig") as fh:
                    for _line in fh:
                        total_rows += 1
                if total_rows > 1:
                    total_rows -= 1  # 去掉表头
            except OSError:
                total_rows = 0

            # P2-4：旧式 progress(int) 保持原节奏（每 10000 行一次，百分 = total//1000 clamp 99）
            # 以兼容旧消费者的精确断言；新式 unified_progress 走 ProgressTracker 阶段权重 + ETA。
            tracker = ProgressTracker(
                stages=[
                    StageSpec("process", weight=8.0, display_name=_("写入 Excel"), has_items=True),
                    StageSpec("save_file", weight=2.0, display_name=_("保存文件")),
                ],
                on_event=self.unified_progress.emit,
            )
            tracker.start()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            sheet_row = 1
            max_rows_per_sheet = 500000
            total = 0

            tracker.begin_stage("process", expected_items=total_rows or 1)
            with open(self._filepath, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                if headers:
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=h)
                    sheet_row = 2

                for row in reader:
                    if self.isInterruptionRequested():
                        tracker.cancel()
                        return
                    total += 1
                    if sheet_row > max_rows_per_sheet:
                        sheet_name = f"Sheet{len(wb.sheetnames) + 1}"
                        ws = wb.create_sheet(sheet_name)
                        sheet_row = 1
                        assert headers is not None
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=h)
                        sheet_row = 2
                    for col, val in enumerate(row, 1):
                        ws.cell(row=sheet_row, column=col, value=excel_safe(val))
                    sheet_row += 1
                    if total % 1000 == 0:
                        tracker.set_item_progress(total)
                    # P2-4：旧式 progress(int) 按 10000 行粒度发出（兼容旧消费方精确断言）
                    if total % 10000 == 0:
                        self.progress.emit(min(total // 1000, 99))
            tracker.end_stage("process")

            if self.isInterruptionRequested():
                tracker.cancel()
                return
            tracker.begin_stage("save_file")
            wb.save(str(self._output_path))
            tracker.finish()
            if not self.isInterruptionRequested():
                self.finished_signal.emit(True, str(self._output_path))
        except Exception as e:  # noqa: BLE001
            if not self.isInterruptionRequested():
                self.finished_signal.emit(False, str(e))


class ResultTable(QWidget):
    """结果查看页 — CSV 表格 + 分页 + 导出。

    CSV 索引与 JSONL 证据查找均通过后台线程执行，避免大文件阻塞 UI。
    """

    # 用户请求在证据查看器中打开某条记录
    record_selected_for_review = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._model = CsvStreamModel(self)
        self._model.indexing_started.connect(self._on_indexing_started)
        self._model.indexing_finished.connect(self._on_indexing_finished)
        self._model.indexing_failed.connect(self._on_indexing_failed)

        self._proxy = QSortFilterProxyModel(self)
        self._proxy.setSourceModel(self._model)
        self._proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._proxy.setFilterKeyColumn(-1)
        self._current_page = 0
        self._export_thread: ExportThread | None = None
        self._evidence_worker: JsonlSearchWorker | None = None
        self._current_evidence_record: dict | None = None  # 当前选中的 JSONL 记录

        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)

        # 信息栏
        info_layout = QHBoxLayout()
        self._info_label = QLabel(_("未加载数据"))
        info_layout.addWidget(self._info_label)
        info_layout.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText(_("搜索当前页全部字段…"))
        self._search.setClearButtonEnabled(True)
        self._search.textChanged.connect(self._apply_filter)
        info_layout.addWidget(self._search)

        # 刷新按钮
        refresh_btn = QPushButton(_("手动刷新"))
        refresh_btn.clicked.connect(self.refresh)
        info_layout.addWidget(refresh_btn)

        # 导出按钮
        export_btn = QPushButton(_("导出 Excel"))
        export_btn.clicked.connect(self._export_excel)
        info_layout.addWidget(export_btn)

        export_md_btn = QPushButton(_("导出 Markdown"))
        export_md_btn.setToolTip(_("将完整抓取结果导出为 Markdown 文件"))
        export_md_btn.clicked.connect(self._export_markdown)
        info_layout.addWidget(export_md_btn)

        export_filtered_btn = QPushButton(_("导出筛选结果"))
        export_filtered_btn.clicked.connect(self._export_filtered_csv)
        info_layout.addWidget(export_filtered_btn)

        # 打开文件夹按钮
        open_folder_btn = QPushButton(_("打开文件夹"))
        open_folder_btn.clicked.connect(self._open_folder)
        info_layout.addWidget(open_folder_btn)

        layout.addLayout(info_layout)

        # 异步加载进度条
        self._loading_bar = QProgressBar()
        self._loading_bar.setRange(0, 0)  # 不确定模式
        self._loading_bar.setTextVisible(False)
        self._loading_bar.setVisible(False)
        self._loading_bar.setMaximumHeight(4)
        layout.addWidget(self._loading_bar)

        # 表格
        self._table = QTableView()
        self._table.setModel(self._proxy)
        header = self._table.horizontalHeader()
        assert header is not None
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(False)  # 流式加载不支持全局排序
        self._evidence = QPlainTextEdit()
        self._evidence.setReadOnly(True)
        self._evidence.setPlaceholderText(_("选择一条记录后，这里显示原始数据、字段证据和质量信息。"))

        # 证据面板容器：证据文本 + 打开按钮
        evidence_container = QWidget()
        evidence_container_layout = QVBoxLayout(evidence_container)
        evidence_container_layout.setContentsMargins(0, 0, 0, 0)
        evidence_container_layout.setSpacing(6)
        evidence_container_layout.addWidget(self._evidence, 1)

        self._open_evidence_btn = QPushButton(_("在证据查看器中打开 →"))
        self._open_evidence_btn.setEnabled(False)
        self._open_evidence_btn.clicked.connect(self._open_in_evidence_view)
        evidence_container_layout.addWidget(self._open_evidence_btn)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self._table)
        splitter.addWidget(evidence_container)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 2)
        layout.addWidget(splitter)
        selection_model = self._table.selectionModel()
        assert selection_model is not None
        selection_model.currentRowChanged.connect(self._show_evidence)

        # 分页控件
        pagination_layout = QHBoxLayout()

        self._first_btn = QPushButton("<<")
        self._first_btn.clicked.connect(self._go_first)
        pagination_layout.addWidget(self._first_btn)

        self._prev_btn = QPushButton("<")
        self._prev_btn.clicked.connect(self._go_prev)
        pagination_layout.addWidget(self._prev_btn)

        self._page_input = QSpinBox()
        self._page_input.setMinimum(1)
        self._page_input.setMaximum(1)
        self._page_input.valueChanged.connect(self._go_to_page)
        pagination_layout.addWidget(self._page_input)

        self._total_label = QLabel("/ 1")
        pagination_layout.addWidget(self._total_label)

        self._next_btn = QPushButton(">")
        self._next_btn.clicked.connect(self._go_next)
        pagination_layout.addWidget(self._next_btn)

        self._last_btn = QPushButton(">>")
        self._last_btn.clicked.connect(self._go_last)
        pagination_layout.addWidget(self._last_btn)

        pagination_layout.addStretch()
        layout.addLayout(pagination_layout)

        self._filepath: Path | None = None
        self._evidence_cache: dict[str, dict] = {}
        self._showing_data = False

    # ---- 公共 API ----

    def load_csv(self, filepath: Path) -> bool:
        """异步加载 CSV 文件。

        立即返回 True 表示已派发索引任务。
        表头与行计数到位后通过 indexing_finished 信号回调更新 UI。
        """
        self._filepath = filepath
        if not filepath.is_file():
            self._info_label.setText(_("未找到结果文件，请确认爬虫已正确输出。点击【手动刷新】重试。"))
            return False

        self._current_page = 0
        self._info_label.setText(_("正在加载结果…"))
        self._model.load_file_async(filepath)
        return True

    # ---- 异步回调（在主线程执行） ----

    def _on_indexing_started(self) -> None:
        """索引任务开始：显示加载指示。"""
        self._loading_bar.setVisible(True)
        self._info_label.setText(_("正在索引 CSV 行…"))
        self._showing_data = False

    def _on_indexing_finished(self, total_rows: int) -> None:
        """索引任务完成：更新信息栏与分页控件。"""
        self._loading_bar.setVisible(False)
        total = self._model.total_rows
        self._info_label.setText(
            _("总行数（近似）: {:,} 行").format(total)
        )
        self._update_pagination()
        self._table.resizeColumnsToContents()
        if total_rows > 0 and not self._showing_data:
            self._fade_in_widget(self._table)
            self._showing_data = True

    def _fade_in_widget(self, widget: QWidget) -> None:
        """对指定控件播放 200ms 淡入动画。"""
        effect = QGraphicsOpacityEffect(widget)
        effect.setOpacity(0.0)
        widget.setGraphicsEffect(effect)
        anim = QPropertyAnimation(effect, b"opacity", widget)
        anim.setDuration(200)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        anim.start(QPropertyAnimation.DeletionPolicy.DeleteWhenStopped)

    def _on_indexing_failed(self, message: str) -> None:
        """索引任务失败：提示错误。"""
        self._loading_bar.setVisible(False)
        self._info_label.setText(_("加载失败：{0}").format(message))

    def _apply_filter(self, text: str) -> None:
        self._proxy.setFilterRegularExpression(QRegularExpression(QRegularExpression.escape(text)))
        self._info_label.setText(
            _("当前页显示 {0} / {1} 行").format(self._proxy.rowCount(), self._model.rowCount())
        )

    def _show_evidence(self, proxy_index: QModelIndex, _previous: QModelIndex) -> None:
        if not proxy_index.isValid() or not self._filepath:
            return
        source_index = self._proxy.mapToSource(proxy_index)
        try:
            record_column = self._model.headers.index("record_id")
        except ValueError:
            self._evidence.setPlainText(_("当前 CSV 不包含 record_id，无法关联字段证据。"))
            self._open_evidence_btn.setEnabled(False)
            self._current_evidence_record = None
            return
        record_id = str(self._model.data(self._model.index(source_index.row(), record_column)) or "")
        if not record_id:
            self._open_evidence_btn.setEnabled(False)
            self._current_evidence_record = None
            return

        # 命中缓存则立即展示
        record = self._evidence_cache.get(record_id)
        if record is not None:
            self._evidence.setPlainText(json.dumps(record, ensure_ascii=False, indent=2, default=str))
            self._current_evidence_record = record
            self._open_evidence_btn.setEnabled(True)
            return

        # 异步查找时暂时禁用按钮
        self._open_evidence_btn.setEnabled(False)
        self._current_evidence_record = None

        jsonl = self._filepath.with_name("records.jsonl")
        if not jsonl.is_file():
            self._evidence.setPlainText(_(f"记录 {record_id}\n未找到配套 records.jsonl 证据文件。"))
            return

        # 异步查找证据，避免大 JSONL 阻塞 UI
        self._evidence.setPlaceholderText(_(f"正在查找记录 {record_id} 的证据…"))
        if self._evidence_worker is not None and self._evidence_worker.isRunning():
            self._evidence_worker.requestInterruption()
            self._evidence_worker.quit()
            self._evidence_worker.wait(1000)

        self._evidence_worker = JsonlSearchWorker(jsonl, record_id, parent=self)
        self._evidence_worker.found.connect(self._on_evidence_found)
        self._evidence_worker.not_found.connect(self._on_evidence_not_found)
        self._evidence_worker.failed.connect(self._on_evidence_failed)
        self._evidence_worker.finished.connect(self._on_evidence_worker_finished)
        self._evidence_worker.start()

    def _on_evidence_found(self, record_id: str, record: dict) -> None:
        """证据查找成功回调（主线程）。"""
        self._evidence_cache[record_id] = record
        self._current_evidence_record = record
        self._evidence.setPlaceholderText(_("选择一条记录后，这里显示原始数据、字段证据和质量信息。"))
        self._evidence.setPlainText(json.dumps(record, ensure_ascii=False, indent=2, default=str))
        self._open_evidence_btn.setEnabled(True)

    def _on_evidence_not_found(self, record_id: str) -> None:
        """证据查找未命中回调（主线程）。"""
        self._current_evidence_record = None
        self._open_evidence_btn.setEnabled(False)
        self._evidence.setPlaceholderText(_("选择一条记录后，这里显示原始数据、字段证据和质量信息。"))
        self._evidence.setPlainText(_(f"记录 {record_id}\n未找到配套 records.jsonl 证据文件。"))

    def _on_evidence_failed(self, message: str) -> None:
        """证据查找失败回调（主线程）。"""
        self._current_evidence_record = None
        self._open_evidence_btn.setEnabled(False)
        self._evidence.setPlaceholderText(_("选择一条记录后，这里显示原始数据、字段证据和质量信息。"))
        self._evidence.setPlainText(_(f"证据加载失败：{message}"))

    def _on_evidence_worker_finished(self) -> None:
        """线程结束后清理引用。"""
        sender = self.sender()
        if sender is self._evidence_worker:
            self._evidence_worker = None

    def _open_in_evidence_view(self) -> None:
        """触发信号：在证据查看器中打开当前记录。"""
        if self._current_evidence_record is not None:
            self.record_selected_for_review.emit(self._current_evidence_record)

    def _export_filtered_csv(self) -> None:
        if self._proxy.rowCount() == 0:
            QMessageBox.information(self, _("导出筛选结果"), _("当前没有可导出的筛选结果。"))
            return
        output_path, _selected_filter = QFileDialog.getSaveFileName(
            self, _("导出筛选结果"), "filtered_records.csv", _("CSV 文件 (*.csv)")
        )
        if not output_path:
            return
        with Path(output_path).open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(self._model.headers)
            for row in range(self._proxy.rowCount()):
                writer.writerow(
                    [self._proxy.data(self._proxy.index(row, column)) for column in range(self._proxy.columnCount())]
                )
        QMessageBox.information(self, _("导出完成"), _(f"已导出 {self._proxy.rowCount()} 行到：\n{output_path}"))

    def _export_markdown(self) -> None:
        """导出完整结果为 Markdown 文件。"""
        if not self._filepath or not self._filepath.is_file():
            QMessageBox.information(self, _("提示"), _("请先加载 CSV 文件"))
            return

        output_path, _selected_filter = QFileDialog.getSaveFileName(
            self, _("导出 Markdown"), self._filepath.stem + ".md",
            _("Markdown 文件 (*.md)"),
        )
        if not output_path:
            return

        # S3.1.1：Markdown 导出移入后台线程（大表不冻结界面）
        from ..core.background_worker import BackgroundWorker, run_worker

        filepath = self._filepath
        jsonl = filepath.with_name("records.jsonl")
        target = Path(output_path)

        class _MarkdownExportWorker(BackgroundWorker):
            def work(self) -> str:
                from omnicrawler.export.markdown_exporter import MarkdownExporter

                MarkdownExporter.export_results(
                    csv_path=filepath,
                    jsonl_path=jsonl if jsonl.is_file() else None,
                    output_path=target,
                    include_evidence=True,
                )
                return str(target)

        run_worker(
            _MarkdownExportWorker(),
            on_succeeded=lambda path: QMessageBox.information(
                self, _("导出成功"), _(f"已导出到: {path}")
            ),
            on_failed=lambda error: QMessageBox.critical(self, _("导出失败"), error),
        )

    def refresh(self) -> None:
        """手动刷新。"""
        if self._filepath and self._filepath.is_file():
            self.load_csv(self._filepath)

    # ---- 分页 ----

    def _update_pagination(self) -> None:
        """更新分页控件状态。"""
        total_pages = self._model.total_pages
        current = self._current_page + 1  # 1-based display

        self._page_input.blockSignals(True)
        self._page_input.setMinimum(1)
        self._page_input.setMaximum(max(1, total_pages))
        self._page_input.setValue(current)
        self._page_input.blockSignals(False)

        self._total_label.setText(f"/ {max(1, total_pages)}")

        self._first_btn.setEnabled(self._current_page > 0)
        self._prev_btn.setEnabled(self._current_page > 0)
        self._next_btn.setEnabled(self._current_page < total_pages - 1)
        self._last_btn.setEnabled(self._current_page < total_pages - 1)

    def _warn_if_page_error(self) -> None:
        """A17：翻页加载失败时给用户可见提示（不再静默）。"""
        if self._model.last_error:
            from ..widgets.toast import ToastManager
            ToastManager.instance().warning(_("翻页失败: {0}").format(self._model.last_error))
            self._model.last_error = None

    def _go_first(self) -> None:
        if self._current_page > 0:
            self._current_page = 0
            self._model.go_to_page(self._current_page)
            self._update_pagination()
            self._table.resizeColumnsToContents()
            self._warn_if_page_error()

    def _go_prev(self) -> None:
        if self._current_page > 0:
            self._current_page -= 1
            self._model.go_to_page(self._current_page)
            self._update_pagination()
            self._table.resizeColumnsToContents()
            self._warn_if_page_error()

    def _go_next(self) -> None:
        if self._current_page < self._model.total_pages - 1:
            self._current_page += 1
            self._model.go_to_page(self._current_page)
            self._update_pagination()
            self._table.resizeColumnsToContents()
            self._warn_if_page_error()

    def _go_last(self) -> None:
        last = max(0, self._model.total_pages - 1)
        if self._current_page != last:
            self._current_page = last
            self._model.go_to_page(self._current_page)
            self._update_pagination()
            self._table.resizeColumnsToContents()
            self._warn_if_page_error()

    def _go_to_page(self, page: int) -> None:
        """页码输入跳转。"""
        # 已通过 setMaximum 限制范围
        actual = page - 1  # 转为 0-based
        if actual != self._current_page and 0 <= actual < self._model.total_pages:
            self._current_page = actual
            self._model.go_to_page(self._current_page)
            self._update_pagination()
            self._table.resizeColumnsToContents()

    # ---- 导出 ----

    def _export_excel(self) -> None:
        """导出为 Excel。"""
        if not self._filepath or not self._filepath.is_file():
            QMessageBox.information(self, _("提示"), _("请先加载 CSV 文件"))
            return

        output_path, _selected_filter = QFileDialog.getSaveFileName(
            self, _("导出 Excel"), self._filepath.stem + ".xlsx",
            _("Excel 文件 (*.xlsx)"),
        )
        if not output_path:
            return

        progress = QProgressDialog(_("正在导出 Excel..."), _("取消"), 0, 0, self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()

        self._export_thread = ExportThread(self._filepath, Path(output_path))
        self._export_thread.setParent(self)
        # S3.1.9：取消按钮真正中断导出线程
        progress.canceled.connect(self._export_thread.requestInterruption)
        self._export_thread.finished_signal.connect(
            lambda ok, msg: self._on_export_finished(ok, msg, progress)
        )
        self._export_thread.start()

    def _on_export_finished(self, ok: bool, msg: str, progress: QProgressDialog) -> None:
        """导出完成处理。"""
        progress.close()
        if ok:
            QMessageBox.information(self, _("导出成功"), _(f"已导出到: {msg}"))
        else:
            QMessageBox.critical(self, _("导出失败"), msg)

    def _open_folder(self) -> None:
        """打开结果文件夹。"""
        from PySide6.QtCore import QUrl
        from PySide6.QtGui import QDesktopServices

        if self._filepath and self._filepath.parent.is_dir():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._filepath.parent)))
        else:
            QMessageBox.information(self, _("提示"), _("未找到结果目录"))
