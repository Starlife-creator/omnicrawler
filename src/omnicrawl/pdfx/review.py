from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import ProjectConfig
from .database import Database
from .utils import utcnow

ACCEPT = {"确认", "已确认", "通过", "接受", "accept", "accepted", "yes", "1"}
REJECT = {"排除", "拒绝", "删除", "reject", "rejected", "delete", "no", "0"}


def _rows_from_csv(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        yield from csv.DictReader(handle)


def _rows_from_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "人工复核队列" not in workbook.sheetnames:
        raise ValueError("Excel中缺少“人工复核队列”工作表")
    sheet = workbook["人工复核队列"]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return
    for row in rows:
        yield {headers[index]: value for index, value in enumerate(row) if index < len(headers)}


def apply_review(config: ProjectConfig, db: Database, file_path: str | Path) -> dict[str, int]:
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"复核文件不存在: {path}")
    rows = _rows_from_xlsx(path) if path.suffix.lower() == ".xlsx" else _rows_from_csv(path)
    summary = {"rows": 0, "accepted": 0, "rejected": 0, "skipped": 0, "missing": 0}
    now = utcnow()
    for row in rows:
        summary["rows"] += 1
        record_id = str(row.get("记录ID") or "").strip().lstrip("'")
        decision = str(row.get("复核决定") or "").strip().casefold()
        if not record_id or not decision:
            summary["skipped"] += 1
            continue
        exists = db.fetchone("SELECT record_id FROM records WHERE record_id=?", (record_id,))
        if not exists:
            summary["missing"] += 1
            continue
        if decision in REJECT:
            db.execute(
                "UPDATE records SET review_status='human_rejected', validation_messages=?, updated_at=? WHERE record_id=?",
                (str(row.get("复核备注") or ""), now, record_id),
            )
            summary["rejected"] += 1
            continue
        if decision not in ACCEPT:
            summary["skipped"] += 1
            continue
        with db.transaction() as conn:
            for spec in config.fields:
                normalized = row.get(spec.label)
                raw = row.get(f"{spec.label}_原始值")
                unit = row.get(f"{spec.label}_单位")
                page = row.get(f"{spec.label}_页码")
                evidence = row.get(f"{spec.label}_原文证据")
                if all(value in {None, ""} for value in (normalized, raw, unit, page, evidence)):
                    continue
                try:
                    page_no = int(float(page)) if page not in {None, ""} else None
                except (TypeError, ValueError):
                    page_no = None
                conn.execute(
                    """
                    INSERT INTO field_values
                        (record_id, field_name, raw_value, normalized_value, unit, page_no,
                         evidence, extraction_method, confidence, validation_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'human_review', 1.0, 'human_reviewed')
                    ON CONFLICT(record_id, field_name) DO UPDATE SET
                        raw_value=excluded.raw_value,
                        normalized_value=excluded.normalized_value,
                        unit=excluded.unit,
                        page_no=excluded.page_no,
                        evidence=excluded.evidence,
                        extraction_method='human_review',
                        confidence=1.0,
                        validation_status='human_reviewed'
                    """,
                    (
                        record_id, spec.name,
                        None if raw is None else str(raw).lstrip("'"),
                        None if normalized is None else str(normalized).lstrip("'"),
                        None if unit is None else str(unit).lstrip("'"),
                        page_no,
                        None if evidence is None else str(evidence).lstrip("'"),
                    ),
                )
            conn.execute(
                """
                UPDATE records SET review_status='human_accepted', validation_status='human_reviewed',
                    validation_messages=?, confidence=1.0, updated_at=? WHERE record_id=?
                """,
                (str(row.get("复核备注") or ""), now, record_id),
            )
        summary["accepted"] += 1
    return summary


