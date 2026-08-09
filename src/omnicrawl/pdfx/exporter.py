from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from .config import ProjectConfig
from .database import Database
from .utils import atomic_output_path, atomic_write_json, utcnow

BASE_HEADERS = [
    "记录ID", "文档ID", "文件名", "原始路径", "记录序号", "文档类型",
    "抽取方式", "记录置信度", "校验状态", "复核状态", "校验信息",
]
NUMERIC_TEXT = re.compile(
    r"^-?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?%?$"
)


def safe_cell(value: Any, max_length: int = 32700) -> Any:
    """Excel 单元格安全化（S4.4 ③：统一委托 core.utils.excel_safe，消除重复实现）。"""
    from ..core.utils import excel_safe

    return excel_safe(value, max_length=max_length)


def _wide_query(config: ProjectConfig, review_only: bool = False) -> tuple[str, list[str]]:
    expressions: list[str] = []
    params: list[str] = []
    for spec in config.fields:
        for column in ("normalized_value", "raw_value", "unit", "page_no", "evidence", "confidence"):
            expressions.append(f"MAX(CASE WHEN fv.field_name=? THEN fv.{column} END)")
            params.append(spec.name)
    where = "WHERE r.review_status='needs_review'" if review_only else ""
    sql = f"""
        SELECT r.record_id, d.doc_id, d.filename, d.primary_path, r.record_index,
               d.document_type, r.extraction_method, r.confidence,
               r.validation_status, r.review_status, r.validation_messages
               {',' if expressions else ''} {','.join(expressions)}
        FROM records r
        JOIN documents d ON d.doc_id=r.doc_id
        LEFT JOIN field_values fv ON fv.record_id=r.record_id
        {where}
        GROUP BY r.record_id
        ORDER BY d.filename, r.record_index
    """
    return sql, params


def wide_headers(config: ProjectConfig) -> list[str]:
    headers = list(BASE_HEADERS)
    for spec in config.fields:
        headers.extend([
            spec.label,
            f"{spec.label}_原始值",
            f"{spec.label}_单位",
            f"{spec.label}_页码",
            f"{spec.label}_原文证据",
            f"{spec.label}_字段置信度",
        ])
    return headers


def write_query_csv(
    db: Database,
    path: Path,
    headers: list[str],
    sql: str,
    params: Iterable[Any] = (),
) -> int:
    count = 0
    with atomic_output_path(path) as temp_path:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            cursor = db.connection.execute(sql, tuple(params))
            for row in cursor:
                writer.writerow([safe_cell(value, max_length=1_000_000) for value in tuple(row)])
                count += 1
    return count


def _csv_to_sheet(workbook: Workbook, path: Path, title: str) -> None:
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title=title)
    sheet.freeze_panes = "A2"
    column_count = 0
    row_count = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row_index, row in enumerate(reader, start=1):
            if row_index == 1:
                cells = []
                for value in row:
                    cell = WriteOnlyCell(sheet, value=safe_cell(value))
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cells.append(cell)
                sheet.append(cells)
                column_count = max(column_count, len(row))
            else:
                sheet.append([safe_cell(value) for value in row])
                if len(row) > column_count:
                    column_count = len(row)
            row_count += 1
            # D48：超过 Excel 行上限（1048576）时停止写入，完整数据保留在 CSV
            if row_count >= 1_048_576:
                sheet.append([f"[已截断：超过 Excel 1048576 行上限，完整数据见 {path.name}]"])
                break
    # D48：auto_filter.ref 按实际列数计算，字段超 115 时不再筛选错位
    if column_count:
        last_column = get_column_letter(column_count)
        sheet.auto_filter.ref = f"A1:{last_column}{max(row_count, 1)}"
    else:
        sheet.auto_filter.ref = "A1"


def export_stage(config: ProjectConfig, db: Database) -> dict[str, Any]:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    output = config.output_dir

    results_sql, results_params = _wide_query(config, review_only=False)
    review_sql, review_params = _wide_query(config, review_only=True)
    headers = wide_headers(config)
    results_csv = output / "results.csv"
    review_csv = output / "review_queue.csv"
    record_count = write_query_csv(db, results_csv, headers, results_sql, results_params)
    review_headers = [*headers, "复核决定", "复核备注"]
    review_count = write_query_csv(db, review_csv, review_headers, review_sql, review_params)

    long_csv = output / "field_values_long.csv"
    long_headers = [
        "记录ID", "文档ID", "文件名", "字段名", "原始值", "标准化值", "单位",
        "页码", "原文证据", "抽取方式", "字段置信度", "校验状态",
    ]
    long_sql = """
        SELECT fv.record_id, d.doc_id, d.filename, fv.field_name, fv.raw_value,
               fv.normalized_value, fv.unit, fv.page_no, fv.evidence,
               fv.extraction_method, fv.confidence, fv.validation_status
        FROM field_values fv
        JOIN records r ON r.record_id=fv.record_id
        JOIN documents d ON d.doc_id=r.doc_id
        ORDER BY d.filename, r.record_index, fv.field_name
    """
    write_query_csv(db, long_csv, long_headers, long_sql)

    documents_csv = output / "documents.csv"
    document_headers = [
        "文档ID", "文件名", "原始路径", "文件大小", "页数", "状态", "文档类型",
        "原生文字页", "OCR页", "候选页", "错误", "解析器版本", "更新时间",
    ]
    document_sql = """
        SELECT doc_id, filename, primary_path, size_bytes, page_count, status,
               document_type, text_page_count, ocr_page_count, candidate_page_count,
               error, parser_version, updated_at
        FROM documents ORDER BY filename
    """
    document_count = write_query_csv(db, documents_csv, document_headers, document_sql)

    errors_csv = output / "errors.csv"
    error_headers = ["文档ID", "阶段", "错误类型", "错误信息", "可重试", "发生时间"]
    error_sql = "SELECT doc_id, stage, error_type, message, retryable, created_at FROM errors ORDER BY id"
    error_count = write_query_csv(db, errors_csv, error_headers, error_sql)

    workbook_path = output / "extraction_results.xlsx"
    workbook = Workbook(write_only=True)
    _csv_to_sheet(workbook, results_csv, "抽取结果")
    _csv_to_sheet(workbook, review_csv, "人工复核队列")
    _csv_to_sheet(workbook, long_csv, "字段长表")
    _csv_to_sheet(workbook, documents_csv, "文件清单")
    _csv_to_sheet(workbook, errors_csv, "错误记录")
    with atomic_output_path(workbook_path, suffix=".xlsx") as temp_path:
        workbook.save(temp_path)

    status_counts = {
        row["status"]: row["n"]
        for row in db.fetchall("SELECT status, COUNT(*) AS n FROM documents GROUP BY status")
    }
    summary = {
        "project": config.project_name,
        "exported_at": utcnow(),
        "documents": document_count,
        "records": record_count,
        "needs_review": review_count,
        "errors": error_count,
        "document_status": status_counts,
        "files": {
            "excel": str(workbook_path),
            "results_csv": str(results_csv),
            "review_csv": str(review_csv),
            "long_csv": str(long_csv),
        },
    }
    atomic_write_json(output / "summary.json", summary)
    return summary
