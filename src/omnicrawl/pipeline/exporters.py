from __future__ import annotations

import csv
import json
import logging
import re
from typing import Any

from ..core.config import AppConfig
from ..core.utils import excel_safe, utcnow
from ..quality.artifact_integrity import verify_artifacts
from ..quality.error_center import build_error_center
from ..quality.quality_report import build_quality_report
from ..state import StateStore

LOGGER = logging.getLogger("omnicrawl")

# Security: validate column names against a strict whitelist before constructing DDL.
# Only standard SQL identifiers (letters, digits, underscores; must start with a
# letter or underscore) are allowed. This prevents SQL injection through
# data-driven column names in the DuckDB export path.
_VALID_COLUMN_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")


def _validate_column_names(columns: list[str]) -> list[str]:
    """Validate that all column names match the safe identifier whitelist.

    Args:
        columns: Column names extracted from data records.

    Returns:
        The validated column names.

    Raises:
        ValueError: If any column name contains characters outside the safe set.
    """
    for name in columns:
        if not _VALID_COLUMN_RE.match(name):
            raise ValueError(
                f"Invalid DuckDB column name: {name!r}. "
                f"Only alphanumeric characters and underscores are allowed "
                f"(must start with a letter or underscore)."
            )
    return columns


def _infer_column_type(name: str, records: list[dict[str, Any]]) -> str:
    """S3.4.1 ③：按样例推断 DuckDB 列类型（不全 VARCHAR）。"""
    kinds: set[str] = set()
    for row in records:
        value = row.get(name)
        if value is None:
            continue
        if isinstance(value, bool):
            kinds.add("bool")
        elif isinstance(value, int):
            kinds.add("int")
        elif isinstance(value, float):
            kinds.add("float")
        else:
            return "VARCHAR"
    if not kinds:
        return "VARCHAR"  # 全 NULL——无类型证据，保守 VARCHAR
    if kinds <= {"int"}:
        return "INTEGER"
    if kinds <= {"int", "float"}:
        return "DOUBLE"
    if kinds <= {"bool"}:
        return "BOOLEAN"
    return "VARCHAR"


def _flatten(prefix: str, value: Any, output: dict[str, Any]) -> None:
    if isinstance(value, dict):
        for key, item in value.items():
            _flatten(f"{prefix}.{key}" if prefix else str(key), item, output)
    elif isinstance(value, (list, tuple)):
        output[prefix] = json.dumps(value, ensure_ascii=False)
    else:
        output[prefix] = value


def _excel_cell(value: Any) -> Any:
    value = excel_safe(value)
    return value[:32700] if isinstance(value, str) else value


def export_all(config: AppConfig, state: StateStore, run_id: str | None = None) -> dict[str, Any]:
    output = config.workspace / "output"
    output.mkdir(parents=True, exist_ok=True)
    where, params = (" WHERE run_id=?", (run_id,)) if run_id else ("", ())
    raw_records = state.rows(f"SELECT * FROM records{where} ORDER BY created_at, record_id", params)
    records: list[dict[str, Any]] = []
    for row in raw_records:
        flat = {
            "record_id": row["record_id"], "source_url": row["source_url"],
            "record_type": row["record_type"], "created_at": row["created_at"],
        }
        _flatten("", json.loads(row["data_json"]), flat)
        records.append(flat)

    files: dict[str, str] = {}
    optional_warnings: list[str] = []
    outputs = config.section("outputs")
    if outputs.get("jsonl", True):
        path = output / "records.jsonl"
        with path.open("w", encoding="utf-8") as handle:
            for row in raw_records:
                # S3.4.1 ①：同一份数据不写两遍——剔除 data_json/evidence_json
                # 原始列，只保留展开的 data/evidence
                clean = {key: value for key, value in row.items() if key not in {"data_json", "evidence_json"}}
                handle.write(json.dumps(
                    {**clean, "data": json.loads(row["data_json"]), "evidence": json.loads(row["evidence_json"])},
                    ensure_ascii=False,
                ) + "\n")
        files["jsonl"] = str(path)
    # S3.4.1 ②：CSV 列按字段定义顺序（extract.fields 键序），附加列按首次出现顺序
    extract_fields = config.section("extract").get("fields", {})
    headers = list(dict.fromkeys([
        "record_id", "source_url", "record_type", "created_at",
        *(str(key) for key in extract_fields if isinstance(extract_fields, dict)),
    ]))
    for row in records:
        for key in row:
            if key not in headers:
                headers.append(key)
    if outputs.get("csv", True):
        path = output / "records.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            # S3.4.1 ④：generator 流式写出，避免二次整表拷贝
            writer.writerows(
                {key: excel_safe(value) for key, value in row.items()} for row in records
            )
        files["csv"] = str(path)

    # S3.4.1 ⑤：responses.csv/errors.csv 文件受 outputs.csv 开关约束；
    # 数据本身始终加载（xlsx 内嵌"抓取清单/错误记录"sheet 也需要）
    response_rows = state.rows(
        f"SELECT final_url, status_code, content_type, size_bytes, content_sha256, raw_path, changed, elapsed_seconds, fetched_at FROM responses{where} ORDER BY id",
        params,
    )
    response_headers = list(response_rows[0]) if response_rows else ["final_url", "status_code", "content_type", "size_bytes", "content_sha256", "raw_path", "changed", "elapsed_seconds", "fetched_at"]
    error_rows = state.rows(f"SELECT url, stage, error_type, message, retryable, created_at FROM errors{where} ORDER BY id", params)
    error_headers = list(error_rows[0]) if error_rows else ["url", "stage", "error_type", "message", "retryable", "created_at"]
    if outputs.get("csv", True):
        response_path = output / "responses.csv"
        with response_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=response_headers)
            writer.writeheader()
            writer.writerows({key: excel_safe(value) for key, value in row.items()} for row in response_rows)
        files["responses_csv"] = str(response_path)

        error_path = output / "errors.csv"
        with error_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=error_headers)
            writer.writeheader()
            writer.writerows({key: excel_safe(value) for key, value in row.items()} for row in error_rows)
        files["errors_csv"] = str(error_path)

    if outputs.get("xlsx", True):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # S2.5.24：xlsx 缺 openpyxl 显式告警（与 parquet/duckdb 一致），不再静默丢弃
            LOGGER.warning(
                "xlsx 导出已启用但未安装 openpyxl；跳过 xlsx（pip install openpyxl）"
            )
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "结构化记录"
            sheet.freeze_panes = "A2"
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            # S3.4.1 ⑦：xlsx 行上限保护（Excel 单表 1,048,576 行硬限）
            XLSX_ROW_LIMIT = 1_000_000
            if len(records) > XLSX_ROW_LIMIT:
                LOGGER.warning(
                    "xlsx 记录数 %d 超过单表上限 %d，超出部分已截断", len(records), XLSX_ROW_LIMIT
                )
            for row in records[:XLSX_ROW_LIMIT]:
                sheet.append([_excel_cell(row.get(key, "")) for key in headers])
            response_sheet = workbook.create_sheet("抓取清单")
            response_sheet.append(response_headers)
            for row in response_rows[:XLSX_ROW_LIMIT]:
                response_sheet.append([_excel_cell(row.get(key, "")) for key in response_headers])
            error_sheet = workbook.create_sheet("错误记录")
            error_sheet.append(error_headers)
            for row in error_rows[:XLSX_ROW_LIMIT]:
                error_sheet.append([_excel_cell(row.get(key, "")) for key in error_headers])
            path = output / "extraction_results.xlsx"
            try:
                workbook.save(path)
            except (PermissionError, OSError) as exc:
                # S3.4.1 ⑧：文件被占用/不可写时友好错误，不再裸抛
                raise RuntimeError(
                    f"无法写入 Excel 文件 {path}——文件可能被其他程序占用或目录不可写: {exc}"
                ) from exc
            files["xlsx"] = str(path)

    if outputs.get("parquet", False):
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            optional_warnings.append(
                "Parquet output was requested but pyarrow is unavailable; install omnicrawl-platform[storage]"
            )
        else:
            # S3.4.1 ③：parquet 保留原始类型（不再全 str 化），pyarrow 自动推断
            table = pa.Table.from_pylist(records) if records else pa.table({"record_id": pa.array([], type=pa.string())})
            path = output / "records.parquet"
            pq.write_table(table, path, compression="zstd")
            files["parquet"] = str(path)

    if outputs.get("duckdb", False):
        try:
            import duckdb
        except ImportError:
            optional_warnings.append(
                "DuckDB output was requested but duckdb is unavailable; install omnicrawl-platform[storage]"
            )
        else:
            path = output / "analytics.duckdb"
            columns = _validate_column_names(headers or ["record_id"])
            # S3.4.1 ③：按样例推断类型（INTEGER/DOUBLE/BOOLEAN/VARCHAR），不全 VARCHAR
            typed = ", ".join(
                f'"{name.replace(chr(34), chr(34) * 2)}" {_infer_column_type(name, records)}'
                for name in columns
            )
            placeholders = ", ".join("?" for _ in columns)
            connection = duckdb.connect(str(path))
            try:
                connection.execute("DROP TABLE IF EXISTS records")
                connection.execute(f"CREATE TABLE records ({typed})")
                if records:
                    # S3.4.1 ③：保留原始类型插入（不再全 str 化）
                    connection.executemany(
                        f"INSERT INTO records VALUES ({placeholders})",
                        [[row.get(key) for key in columns] for row in records],
                    )
            finally:
                connection.close()
            files["duckdb"] = str(path)

    quality_report = build_quality_report(config, state, run_id)
    files["quality_report_json"] = quality_report["files"]["json"]
    files["quality_report_html"] = quality_report["files"]["html"]
    error_center = build_error_center(state, output, run_id)
    files["error_center_json"] = error_center["files"]["json"]
    files["error_center_html"] = error_center["files"]["html"]
    artifact_integrity = verify_artifacts(state, run_id)
    integrity_path = output / "artifact_integrity.json"
    integrity_path.write_text(
        json.dumps(artifact_integrity, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    files["artifact_integrity"] = str(integrity_path)
    summary = {
        "project": config.project_name, "run_id": run_id, "exported_at": utcnow(),
        "records": len(records), "responses": len(response_rows), "errors": len(error_rows),
        "files": files, "warnings": optional_warnings, "quality": quality_report,
        "error_center": error_center,
        "artifact_integrity": artifact_integrity,
    }
    summary_path = output / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    files["summary"] = str(summary_path)
    return summary


def register(registry) -> None:
    registry.register_exporter("default", export_all)
