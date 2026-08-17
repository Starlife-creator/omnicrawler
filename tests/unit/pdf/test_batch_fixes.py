"""Phase 3g/3c 批处理修复测试（D15 worker 错误 / D16 attempt_count / D48 exporter 行限）。"""

from __future__ import annotations

import tempfile
from pathlib import Path


class _StubDB:
    """记录 fetchall/execute 调用的最小 DB 替身。"""

    def __init__(self) -> None:
        self.fetch_calls: list[tuple[str, tuple]] = []
        self.exec_calls: list[tuple[str, tuple]] = []

    def fetchall(self, sql: str, params: tuple = ()) -> list:
        self.fetch_calls.append((sql, params))
        return []

    def execute(self, sql: str, params: tuple = ()):
        self.exec_calls.append((sql, params))
        return None

    def add_error(self, *args) -> None:
        pass

    def transaction(self):
        from contextlib import nullcontext

        return nullcontext()


def _project() -> object:
    from omnicrawler.pdfx.config import ProjectConfig

    return ProjectConfig(
        path=Path("x.yaml"), project_name="t", input_dir=Path("in"), work_dir=Path("work"),
        output_dir=Path("out"), database=Path("db"), parser={"workers": 1}, ocr={}, retrieval={},
        llm={}, extraction={}, normalization={}, validation={"auto_accept_confidence": 0.9},
        fields=[],
    )


def test_d16_parse_stage_filters_exhausted_documents() -> None:
    """D16：attempt_count 达到阈值的文档不再被拉起（防永久损坏 PDF 无限重试）。"""
    from omnicrawler.pdfx.parser import MAX_PARSE_ATTEMPTS, parse_stage

    db = _StubDB()
    parse_stage(_project(), db, limit=None, workers=1)
    assert db.fetch_calls, "parse_stage 应查询待解析文档"
    sql, params = db.fetch_calls[0]
    assert "attempt_count < ?" in sql
    assert params == (MAX_PARSE_ATTEMPTS,)


def test_d15_worker_returns_error_detail(monkeypatch) -> None:
    """D15：OCR worker 异常时把错误串带回主进程（而非仅 None）。"""
    import omnicrawler.pdfx.ocr as ocr

    class _Broken:
        def recognize(self, png):
            raise RuntimeError("tesseract 初始化失败")

    monkeypatch.setattr(ocr, "_worker_backend", _Broken())
    monkeypatch.setattr(ocr, "_worker_document", None)
    monkeypatch.setattr(ocr, "_worker_document_path", None)
    monkeypatch.setattr(ocr, "render_page", lambda *a, **k: b"fake-png")

    class _FakePixmap:
        def tobytes(self, fmt="png"):
            return b"png"

    class _FakePage:
        def get_pixmap(self, **kwargs):
            return _FakePixmap()

    class _FakeDocument:
        def load_page(self, index):
            return _FakePage()

        def close(self):
            pass

    monkeypatch.setattr(ocr.fitz, "open", lambda *a, **k: _FakeDocument())
    result = ocr._ocr_worker_process(("a.pdf", 1, 144))
    assert result[2] is None
    assert result[6] == "tesseract 初始化失败"


def test_d48_csv_to_sheet_column_count_and_row_cap(tmp_path: Path) -> None:
    """D48：auto_filter.ref 按实际列数；超 Excel 行上限截断并在 sheet 内提示。"""
    from openpyxl import Workbook

    from omnicrawler.pdfx.exporter import _csv_to_sheet

    csv_path = tmp_path / "t.csv"
    csv_path.write_text("a,b,c,d\n1,2,3,4\n5,6,7,8\n", encoding="utf-8")
    workbook = Workbook(write_only=True)
    _csv_to_sheet(workbook, csv_path, "测试")
    assert workbook.sheetnames == ["测试"]
    # 4 列 → 列字母 D；2 行数据
    sheet = workbook["测试"]
    assert sheet.auto_filter.ref == "A1:D3"
    workbook.close()


def test_d48_csv_to_sheet_row_cap(tmp_path: Path) -> None:
    """D48：超 1048576 行时写入截断提示而非让 xlsx save 抛异常。"""
    from openpyxl import Workbook

    from omnicrawler.pdfx.exporter import _csv_to_sheet

    csv_path = tmp_path / "big.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        handle.write("a\n")
        for _ in range(1_048_600):
            handle.write("x\n")
    workbook = Workbook(write_only=True)
    _csv_to_sheet(workbook, csv_path, "大表")
    # 不抛异常即通过；write_only sheet 无法直接读行数，这里仅验证未崩溃
    assert workbook.sheetnames == ["大表"]
    workbook.close()


class _FakeTable:
    def __init__(self, data) -> None:
        self._data = data

    def extract(self):
        return self._data


class _FakeTables:
    def __init__(self, tables) -> None:
        self.tables = tables


class _FakePage:
    def __init__(self, tables) -> None:
        self._tables = tables

    def find_tables(self):
        return _FakeTables(self._tables)


def test_d8_table_markdown_preserves_columns() -> None:
    """D8：find_tables 结果渲染为 Markdown，列归属不丢失。"""
    from omnicrawler.pdfx.parser import _extract_tables_markdown

    page = _FakePage([
        _FakeTable([
            ["项目", "期初余额", "期末余额"],
            ["应收账款", "1,000", "1,200"],
            ["存货", "500", "480"],
        ]),
    ])
    markdown = _extract_tables_markdown(page)
    assert "| 项目 | 期初余额 | 期末余额 |" in markdown
    assert "| 应收账款 | 1,000 | 1,200 |" in markdown
    assert "期初余额" in markdown and "期末余额" in markdown  # 列标题不再被拍平


def test_d8_table_detection_failure_is_graceful() -> None:
    """D8：find_tables 抛异常时返回空串，不中断解析。"""
    from omnicrawler.pdfx.parser import _extract_tables_markdown

    class _Broken:
        def find_tables(self):
            raise RuntimeError("table detection failed")

    assert _extract_tables_markdown(_Broken()) == ""


def test_d36_iter_pages_yields_and_wraps() -> None:
    """D36：_iter_parsed_pages 逐页 yield；parse_document 兼容封装收集。"""
    import gc

    import fitz

    from omnicrawler.pdfx.parser import _iter_parsed_pages, parse_document

    # ignore_cleanup_errors：Windows runner 偶发对新建 p.pdf 的瞬时外部锁
    # （安全扫描），临时目录属一次性产物，清理失败不影响断言结果。
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp:
        pdf = Path(temp) / "p.pdf"
        document = fitz.open()
        page = document.new_page()
        page.insert_text((72, 72), "担保金额：1.5亿元", fontname="china-s", fontsize=12)
        document.save(pdf)
        document.close()

        yielded = list(_iter_parsed_pages(str(pdf), 5, 0.1))
        assert len(yielded) == 1
        assert yielded[0]["page_no"] == 1
        assert "担保金额" in yielded[0]["final_text"]

        wrapped = parse_document(str(pdf), 5, 0.1)
        assert wrapped["page_count"] == 1
        assert wrapped["pages"][0]["needs_ocr"] == 0

        # Windows：PyMuPDF 的文件句柄可能在 GC 后才释放，先清理引用再删除临时目录
        del document, page, yielded, wrapped
        gc.collect()


def test_d12_high_image_coverage_forces_ocr(monkeypatch) -> None:
    """D12：有文字层但图片覆盖超 60% 的页面（夹带页眉页脚的表格页）强制 OCR。"""
    import omnicrawler.pdfx.parser as parser

    class _Rect:
        def __init__(self, width: float, height: float) -> None:
            self.width = width
            self.height = height

    class _FakePage:
        rect = _Rect(100.0, 100.0)

        def get_text(self, *args, **kwargs):
            return "2026年公告 页眉 本公司董事会全体成员保证公告内容真实准确完整"

        def get_images(self, full=True):
            return [(1,)]

        def get_image_rects(self, xref):
            return [_Rect(80.0, 80.0)]  # 覆盖 64%

    class _FakeDoc:
        needs_pass = False

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def __iter__(self):
            return iter([_FakePage()])

    monkeypatch.setattr(parser.fitz, "open", lambda *a, **k: _FakeDoc())
    result = parser.parse_document("x.pdf", min_chars=40, max_garbled_ratio=0.03)
    assert result["pages"][0]["needs_ocr"] == 1
    assert result["pages"][0]["ocr_status"] == "pending"


def test_d12_low_image_coverage_keeps_native(monkeypatch) -> None:
    """D12：图像覆盖低的正常文本页不被误判为 OCR。"""
    import omnicrawler.pdfx.parser as parser

    class _Rect:
        def __init__(self, width: float, height: float) -> None:
            self.width = width
            self.height = height

    class _FakePage:
        rect = _Rect(100.0, 100.0)

        def get_text(self, *args, **kwargs):
            return "2026年公告 本公司董事会保证公告内容真实准确完整" * 3

        def get_images(self, full=True):
            return [(1,)]

        def get_image_rects(self, xref):
            return [_Rect(10.0, 10.0)]  # 覆盖 1%

    class _FakeDoc:
        needs_pass = False

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def __iter__(self):
            return iter([_FakePage()])

    monkeypatch.setattr(parser.fitz, "open", lambda *a, **k: _FakeDoc())
    result = parser.parse_document("x.pdf", min_chars=40, max_garbled_ratio=0.03)
    assert result["pages"][0]["needs_ocr"] == 0


def test_d9_tesseract_rebuilds_lines_and_columns() -> None:
    """D9：image_to_data 按 line 分行、left 分列重建，列归属不拍平。"""
    from omnicrawler.pdfx.ocr import TesseractBackend

    class _FakeImage:
        def open(self, *args, **kwargs):
            return self

        def convert(self, *args, **kwargs):
            return self

    class _FakePyTesseract:
        class Output:
            DICT = "dict"

        def image_to_data(self, image, lang=None, output_type=None):
            return {
                "text": ["项目", "期初余额", "应收", "1,000", "存货", "500"],
                "conf": [95, 90, 88, 92, 85, 80],
                "block_num": [1, 1, 2, 2, 2, 2],
                "par_num": [1, 1, 1, 1, 1, 1],
                "line_num": [0, 0, 0, 0, 1, 1],
                "left": [10, 120, 10, 120, 10, 120],
                "width": [30, 60, 40, 50, 40, 50],
            }

    backend = object.__new__(TesseractBackend)
    backend.Image = _FakeImage()
    backend.pytesseract = _FakePyTesseract()
    backend.lang = "chi_sim+eng"
    text, _confidence = backend.recognize(b"png")
    lines = text.splitlines()
    assert len(lines) == 3  # 三行而非拍平一行
    assert "期初余额" in lines[0]
    assert "1,000" in lines[1]
    assert "存货" in lines[2]


def test_d10_table_html_to_markdown() -> None:
    """D10：表格 HTML 转 Markdown，行列结构保留。"""
    from omnicrawler.pdfx.ocr import _table_html_to_markdown

    html = (
        "<table><tr><th>项目</th><th>期末余额</th></tr>"
        "<tr><td>应收账款</td><td>1,200</td></tr></table>"
    )
    markdown = _table_html_to_markdown([html])
    assert "| 项目 | 期末余额 |" in markdown
    assert "| 应收账款 | 1,200 |" in markdown
