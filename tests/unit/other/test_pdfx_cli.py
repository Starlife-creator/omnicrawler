from __future__ import annotations

import csv
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest

pytest.importorskip("openpyxl", reason="PDF review XLSX tests require the optional openpyxl dependency")
from openpyxl import Workbook

from omnicrawl.pdfx import cli
from omnicrawl.pdfx.review import _rows_from_xlsx, apply_review


class _Database:
    def __init__(self, _path):
        self.reset = []

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return None

    def reset_stage(self, stage):
        self.reset.append(stage)


def _fake_config(tmp_path: Path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    return SimpleNamespace(
        path=tmp_path / "fields.yaml",
        input_dir=input_dir,
        database=tmp_path / "work" / "pipeline.sqlite3",
        output_dir=tmp_path / "output",
        work_dir=tmp_path / "work",
        ocr={"backend": "none"},
        llm={"provider": "disabled"},
        parser={},
        extraction={},
        retrieval={},
        validation={},
    )


def _main(monkeypatch, args):
    monkeypatch.setattr(sys, "argv", ["pdf-core", *args])
    cli.main()


def test_pdfx_cli_all_offline_stage_branches(tmp_path: Path, monkeypatch, capsys) -> None:
    config = _fake_config(tmp_path)
    calls = []
    monkeypatch.setattr(cli, "load_config", lambda _path: config)
    monkeypatch.setattr(cli, "Database", _Database)
    monkeypatch.setattr(cli, "validate_runtime_config", lambda _config: ["warning"])
    monkeypatch.setattr(cli, "ingest", lambda *_args: calls.append("ingest") or {"new": 1})
    monkeypatch.setattr(cli, "parse_stage", lambda *_args: calls.append("parse") or {"parsed": 1})
    monkeypatch.setattr(cli, "ocr_stage", lambda *a, **kw: calls.append("ocr") or {"pages": 1})
    monkeypatch.setattr(cli, "extraction_stage", lambda *_args: calls.append("extract") or {"records": 1})
    monkeypatch.setattr(cli, "export_stage", lambda *_args: calls.append("export") or {"records": 1})
    monkeypatch.setattr(cli, "export_text_stage", lambda *_args: calls.append("text") or {"pages": 1})
    monkeypatch.setattr(cli, "status", lambda _db: calls.append("status") or {"ok": True})
    monkeypatch.setattr(cli, "apply_review", lambda *_args: calls.append("review") or {"accepted": 1})
    # S2.3.4：run/process 统一走 service 入口（CLI 不再手写阶段链）
    run_calls: list[str] = []
    monkeypatch.setattr(
        cli, "run_extraction",
        lambda *_args, **_kw: run_calls.append("run") or {"status": {}},
    )
    monkeypatch.setattr(
        cli, "run_processing",
        lambda *_args, **_kw: run_calls.append("process") or {"status": {}},
    )

    common = ["--config", str(config.path)]
    for command, options in (
        ("doctor", []),
        ("ingest", ["--limit", "2"]),
        ("parse", ["--limit", "2", "--workers", "1"]),
        ("ocr", ["--limit-pages", "2"]),
        ("extract", ["--limit", "2", "--workers", "1"]),
        ("export", []),
        ("export-text", ["--limit", "2"]),
        ("status", []),
        ("reset", ["parse"]),
        ("apply-review", ["--file", str(tmp_path / "review.csv")]),
        ("run", ["--limit", "2", "--workers", "1"]),
        ("run", ["--limit", "2", "--skip-ocr"]),
        ("process", ["--limit", "2", "--workers", "1"]),
        ("process", ["--limit", "2", "--skip-ocr"]),
    ):
        _main(monkeypatch, [*common, command, *options])
    assert {"ingest", "parse", "ocr", "extract", "export", "text", "status", "review"}.issubset(calls)
    # S2.3.4：run/process 分支必须经 service 入口（各两次：带/不带 --skip-ocr）
    assert run_calls.count("run") == 2
    assert run_calls.count("process") == 2
    assert '"stage": "run"' in capsys.readouterr().out


def test_pdfx_cli_validate_doctor_dependencies_and_errors(tmp_path: Path, monkeypatch, capsys) -> None:
    monkeypatch.setattr(cli, "validate_project_template", lambda _path: {"valid": True})
    _main(monkeypatch, ["--config", str(tmp_path / "fields.yaml"), "validate"])
    monkeypatch.setattr(cli, "validate_project_template", lambda _path: {"valid": False})
    with pytest.raises(SystemExit) as invalid:
        _main(monkeypatch, ["--config", str(tmp_path / "fields.yaml"), "validate"])
    assert invalid.value.code == 2

    config = _fake_config(tmp_path)
    config.ocr = {"backend": "paddle"}
    monkeypatch.setattr("importlib.util.find_spec", lambda _name: None)
    report = cli.doctor(config)
    assert report["dependencies"]["paddleocr"] is False
    assert report["dependencies"]["paddle"] is False
    config.ocr = {"backend": "tesseract"}
    monkeypatch.setattr("shutil.which", lambda _name: None)
    report = cli.doctor(config)
    assert report["dependencies"]["pytesseract"] is False
    assert report["dependencies"]["tesseract_program"] is False

    monkeypatch.setattr(cli, "load_config", lambda _path: (_ for _ in ()).throw(ValueError("bad config")))
    monkeypatch.setattr(sys, "argv", ["pdf-core", "--config", "bad.yaml", "doctor"])
    with pytest.raises(SystemExit) as failed:
        cli.main()
    assert failed.value.code == 1
    assert "bad config" in capsys.readouterr().err


class _Transaction:
    def __init__(self, db):
        self.db = db

    def __enter__(self):
        return self.db

    def __exit__(self, *_args):
        return None


class _ReviewDatabase:
    def __init__(self, existing):
        self.existing = set(existing)
        self.statements = []

    def fetchone(self, _sql, values):
        return {"record_id": values[0]} if values[0] in self.existing else None

    def execute(self, sql, values):
        self.statements.append((" ".join(sql.split()), values))

    def transaction(self):
        return _Transaction(self)


def test_apply_review_csv_accept_reject_missing_skipped_and_field_updates(tmp_path: Path) -> None:
    path = tmp_path / "review.csv"
    headers = [
        "记录ID",
        "复核决定",
        "复核备注",
        "金额",
        "金额_原始值",
        "金额_单位",
        "金额_页码",
        "金额_原文证据",
    ]
    rows = [
        ["'accepted", "确认", "ok", "100", "'1百", "元", "2", "evidence"],
        ["rejected", "拒绝", "bad", "", "", "", "", ""],
        ["missing", "确认", "", "", "", "", "", ""],
        ["accepted", "maybe", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)

    config = SimpleNamespace(fields=[SimpleNamespace(name="amount", label="金额")])
    db = _ReviewDatabase({"accepted", "rejected"})
    summary = apply_review(config, db, path)
    assert summary == {"rows": 5, "accepted": 1, "rejected": 1, "skipped": 2, "missing": 1}
    # 参数顺序为 (record_id, field_name, raw, normalized, unit, page_no, evidence)：
    # page_no（2）应在第 6 位，record_id 在第 1 位（blocking 回归）
    assert any("field_values" in sql and values[0] == "accepted" and values[5] == 2
               for sql, values in db.statements)
    assert any("human_rejected" in sql for sql, _values in db.statements)
    assert any("human_accepted" in sql for sql, _values in db.statements)

    with pytest.raises(FileNotFoundError):
        apply_review(config, db, tmp_path / "missing.csv")


def test_review_xlsx_reader_sheet_validation_empty_and_rows(tmp_path: Path) -> None:
    invalid = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    workbook.save(invalid)
    with pytest.raises(ValueError, match="人工复核队列"):
        list(_rows_from_xlsx(invalid))

    empty = tmp_path / "empty.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工复核队列"
    workbook.save(empty)
    assert list(_rows_from_xlsx(empty)) == []

    valid = tmp_path / "valid.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工复核队列"
    sheet.append(["记录ID", "复核决定"])
    sheet.append(["one", "确认"])
    workbook.save(valid)
    assert list(_rows_from_xlsx(valid)) == [{"记录ID": "one", "复核决定": "确认"}]
